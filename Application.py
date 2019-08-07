import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
from tkinter import *
from PIL import ImageTk, Image
import glob
import openpyxl
import os
from os import listdir
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2TkAgg
from matplotlib.figure import Figure
import Data
import AlgemeneInfo
import WorkbookLayout
import Main

class Application(tk.Frame):

    def __init__(self, master=None):
        """
        initialisation of the gui
        """
        super().__init__(master)
        self.master = master
        self.pack()
        self.create_widgets()

    def create_widgets(self):
        """
        create the different widgets of the gui and put them in rows and columns
        """
        self.grid()
        self.pack(fill="both", expand=True)

        #styles
        bgColor = '#343434'
        fgColor = '#FFFFFF'
        red = '#f91919'
        gui_style = ttk.Style()
        gui_style.map("My.TButton",background=[('pressed', '!disabled', bgColor), ('active', bgColor), ('!active', bgColor)])
        gui_style.configure('My.TFrame', background=bgColor)
        gui_style.configure('My.TNotebook', background=fgColor)
        gui_style.configure('My.TLabel', background=bgColor, foreground=fgColor, font=('verdana', 10, ''))
        gui_style.configure('My.TCheckbutton', background=bgColor, foreground=fgColor, font=('verdana', 10, ''))
        gui_style.configure('My.TLabelframe.Label', background=bgColor, foreground=fgColor)
        gui_style.configure('My.TRadiobutton', background=bgColor, foreground=fgColor, font=('verdana', 10, ''))
        gui_style.configure('My.TSeparator', background='#000000')

        #tabs
        tabControl = ttk.Notebook(self, style='My.TNotebook')
        tabControl.pack(expand=1, fill='both')

        self.tabBifi = ttk.Frame(tabControl, style='My.TFrame')
        tabControl.add(self.tabBifi, text='Processing')

        self.tabInfo = ttk.Frame(tabControl, style='My.TFrame')
        tabControl.add(self.tabInfo, text='Graphs')

        #variables
        self.f = Figure(figsize=(5, 4), dpi=100)
        self.ax = self.f.add_subplot(111)
        self.fb = Figure(figsize=(5, 4), dpi=100)
        self.bx = self.fb.add_subplot(111)
        self.subfolders = []
        self.filePath = ''
        self.samples = []
        self.wbName = ''
        self.hours = []
        # sheetNames = []
        self.sampleNames = []
        # times = []
        self.sampleNumbers = []
        self.sampleMeasures = []
        self.newFile = ''
        self.hourRb = tk.IntVar()
        self.machineRb = tk.IntVar()
        self.sampleCb = []
        self.sampleRb = tk.IntVar()
        self.hourVar = tk.IntVar()
        self.hourVar2 = tk.IntVar()
        self.timeCb = tk.IntVar()
        self.timeInt = tk.IntVar()
        self.fileVar = tk.StringVar()
        self.comboVar = tk.StringVar()
        self.comboList = []     # nr + file
        self.configContent = [] # nr + filepath
        self.samples = []       # frame + filepath
        self.ivCb = tk.IntVar()
        self.eqeCb = tk.IntVar()
        self.photoCb = tk.IntVar()

        self.filePsc = tk.StringVar()
        self.comboPsc = tk.StringVar()
        self.comboListPsc = ['']
        self.configPsc = []
        self.samplesPsc = []

        self.fileSm = tk.StringVar()
        self.comboSm = tk.StringVar()
        self.comboListSm = [''] # nr + configContent
        self.configSm = []      # filepath

        self.rsh = []

        #Processing
            # Machine choise
        ttk.Separator(self.tabBifi, orient=HORIZONTAL, style='My.TSeparator').grid(row=0, columnspan=10, sticky='WE', padx=5)
        self.machineLabel = ttk.Label(self.tabBifi, text='Select device', style='My.TLabel').grid(sticky='W', row=0, column=0, columnspan=2, padx=10)
        self.RbLoana = ttk.Radiobutton(self.tabBifi, text='LOANA', variable=self.machineRb, value=1, style='My.TRadiobutton', command = lambda:self.selectMachine())\
            .grid(row=1,sticky='W', padx=5, pady=(0,5))
        self.RbThin = ttk.Radiobutton(self.tabBifi, text='PME', variable=self.machineRb, value=2, style='My.TRadiobutton', command=lambda: self.selectMachine())\
            .grid(row=1, column=1,sticky='W', pady=(0,5))
        self.machineRb.set(1)
        self.RbSwitch = ttk.Radiobutton(self.tabBifi, text='K2400', variable=self.machineRb, value=3,
                                        style='My.TRadiobutton', command=lambda: self.selectMachine()) \
            .grid(row=1, column=2, sticky='W', pady=(0,5))
        self.machineRb.set(1)
        ttk.Separator(self.tabBifi, orient=HORIZONTAL, style='My.TSeparator').grid(row=2, columnspan=10, sticky='WE', padx=5)
            # Hour choise
        self.hourLabel = ttk.Label(self.tabBifi, text='Stressing duration?', style='My.TLabel').grid(sticky='W', row=2, column=0, columnspan=3, padx=10)
        self.hourRbAll = ttk.Radiobutton(self.tabBifi, text='All', variable=self.hourRb, value=1,style='My.TRadiobutton').grid(row=3, sticky='W', padx=5)
        self.hourRbEntry = ttk.Radiobutton(self.tabBifi, text='Between', variable=self.hourRb, value=2,style='My.TRadiobutton').grid(row=3, column=1,sticky='W')
        self.hourRb.set(1)
        self.hourInput = ttk.Entry(self.tabBifi, textvariable=self.hourVar)
        self.hourInput.grid(sticky='WE', row=3, column=2, pady=5,padx=(5,5))
        self.hourLabel = ttk.Label(self.tabBifi, text='and', style='My.TLabel').grid(sticky='W', row=3,column=3,padx=5)
        self.hourInput2 = ttk.Entry(self.tabBifi, textvariable=self.hourVar2)
        self.hourInput2.grid(sticky='WE', row=3, column=4, pady=5, padx=(5, 5))
        self.timeInterval = ttk.Checkbutton(self.tabBifi, text="Data every", variable=self.timeCb, onvalue=1, offvalue=0,style='My.TCheckbutton')
        self.timeInterval.grid(row=3, column=5, sticky='W', padx=5, pady=5)
        self.timeInput = ttk.Entry(self.tabBifi, textvariable=self.timeInt)
        self.timeInput.grid(sticky='WE', row=3, column=6, pady=5, padx=(5, 5))
        self.everyTimeLabel = ttk.Label(self.tabBifi, text='h', style='My.TLabel')
        self.everyTimeLabel.grid(sticky='W', row=3, column=7, padx=5)
        ttk.Separator(self.tabBifi, orient=HORIZONTAL, style='My.TSeparator').grid(row=4, columnspan=10, sticky='WE', padx=5)
            # Excel file
        self.excelLabel = ttk.Label(self.tabBifi, text='Select Excel file', style='My.TLabel').grid(sticky='W',row=4,column=0,columnspan=2, padx=10)
        self.fileInput = ttk.Entry(self.tabBifi, textvariable=self.fileVar)
        self.fileInput.grid(row=6, column=1, columnspan=8, sticky='WE', padx=5, pady=5)
        self.fileBtn = ttk.Button(self.tabBifi, text='Select file', style='My.TButton')
        self.fileBtn.bind('<ButtonRelease-1>', self.pickFile)
        self.fileBtn.grid(sticky='WE', row=6, column=0, padx=5, pady=5)
            # Data choise
        ttk.Separator(self.tabBifi, orient=HORIZONTAL, style='My.TSeparator').grid(row=7, columnspan=10, sticky='WE', padx=5)
        self.dataLabel = ttk.Label(self.tabBifi, text='Select data types', style='My.TLabel').grid(sticky='W', row=7, column=0,columnspan=2, padx=10)
        self.iv = ttk.Checkbutton(self.tabBifi, text="IV", variable=self.ivCb, onvalue=1, offvalue=0,style='My.TCheckbutton')
        self.iv.grid(row=8, column=0, sticky='W', padx=5, pady=5)
        self.eqe = ttk.Checkbutton(self.tabBifi, text="EQE", variable=self.eqeCb, onvalue=1, offvalue=0,style='My.TCheckbutton')
        self.eqe.grid(row=8, column=1, sticky='W', pady=5)
        self.photo = ttk.Checkbutton(self.tabBifi, text="EL picture", variable=self.photoCb, onvalue=1, offvalue=0,style='My.TCheckbutton')
        self.photo.grid(row=8, column=2, sticky='W', pady=5)
            # Start
        self.beginBtn = ttk.Button(self.tabBifi, text='Begin', command=self.begin, style='My.TButton')
        self.beginBtn.grid(sticky='W',row=9,column=0,padx=5,pady=5)
        self.errorLabel = ttk.Label(self.tabBifi, text='', background=bgColor, foreground=red, font=('verdana', 10, ''))
        self.errorLabel.grid(sticky='W', row=9, column=1, columnspan=3)
        img = Image.open("UHasselt.png")
        basewidth = 120 #70 voor zelfde breedte
        wpercent = (basewidth / float(img.size[0]))
        hsize = int((float(img.size[1]) * float(wpercent)))
        img = img.resize((basewidth, hsize), Image.ANTIALIAS)
        img = ImageTk.PhotoImage(img)
        label1 = Label(self.tabBifi, image=img, background='white')
        label1.image = img
        label1.grid(row=0, column=6, columnspan=2, padx=5, pady=(5,5), sticky='E')

        #configure column 2 to stretch with the window
        self.tabBifi.grid_columnconfigure(7, weight=1)

        #Info
        self.cbFrame = ttk.Frame(self.tabInfo, style='My.TFrame')
        self.cbFrame.grid(row=0, columnspan=5, sticky='W', padx=5)
        self.btn = ttk.Button(self.tabInfo, text='Begin', command=self.graph, style='My.TButton')
        self.btn.grid(sticky='W', row=1, column=0, padx=5, pady=5)
        self.gFrame = ttk.Frame(self.tabInfo, style='My.TFrame')
        self.gFrame.grid(row=2, rowspan=2, sticky='W', padx=5)
        self.gFrame2 = ttk.Frame(self.tabInfo, style='My.TFrame')
        self.gFrame2.grid(row=2, rowspan=2, column=1, sticky='W', padx=5)
        self.tFrame = ttk.Frame(self.tabInfo, style='My.TFrame')
        self.tFrame.grid(row=2, column=2, sticky='W', padx=5)

        self.canvas = FigureCanvasTkAgg(self.f, self.gFrame)
        self.canvas.show()
        self.canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
        self.toolbar = NavigationToolbar2TkAgg(self.canvas, self.gFrame)
        self.toolbar.update()
        self.canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        # Shrink current axis by 25%
        box = self.ax.get_position()
        self.ax.set_position([box.x0, box.y0, box.width * 0.75, box.height])

        self.canvas2 = FigureCanvasTkAgg(self.fb, self.gFrame2)
        self.canvas2.show()
        self.canvas2.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
        self.toolbar = NavigationToolbar2TkAgg(self.canvas2, self.gFrame2)
        self.toolbar.update()
        self.canvas2._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        # Shrink current axis by 25%
        box = self.bx.get_position()
        self.bx.set_position([box.x0, box.y0, box.width * 0.75, box.height])

    def graph(self):
        self.ax.clear()
        self.bx.clear()
        iv = []
        self.f = plt.figure(1)
        if self.machineRb.get() == 1:
            for i in range(0, len(self.samples)):
                if self.sampleCb[i].get() == 1:
                    for h in range(0, len(self.subfolders)):
                        if os.path.exists(self.filePath + '/' + str(self.subfolders[h]) + '/' + self.samples[i] + '/IV/' +
                                          self.samples[i] + '.lgt'):
                            iv = Data.Data.getDataList(
                                self.filePath + '/' + str(self.subfolders[h]) + '/' + self.samples[i] + '/IV/' +
                                self.samples[i] + '.lgt')
                            div = Data.Data.getDataList(
                                self.filePath + '/' + str(self.subfolders[h]) + '/' + self.samples[i] + '/IV/' +
                                self.samples[i] + '.drk')
                            self.ax.plot(iv[1], iv[0], label=str(self.samples[i] + ' - ' + str(self.hours[h])) + 'h')
                            self.bx.plot(div[1], div[0], label=str(self.samples[i] + ' - ' + str(self.hours[h])) + 'h')

                            self.ax.set_xlabel('Voltage [V]')
                            self.ax.set_ylabel('Current [A]')
                            self.ax.set_title('Light IV')
                            # Put a legend to the right of the current axis
                            self.ax.legend(loc='center left', bbox_to_anchor=(1, 0.5))

                            # self.ax.legend()
                            self.bx.set_xlabel('Voltage [V]')
                            self.bx.set_ylabel('Current [A]')
                            self.bx.set_title('Dark IV')
                            self.bx.legend()
                            # Put a legend to the right of the current axis
                            self.bx.legend(loc='center left', bbox_to_anchor=(1, 0.5))

        elif self.machineRb.get() == 2:
            for i in range(0, len(self.samples)):
                if self.sampleCb[i].get() == 1:
                    for h in self.hours:
                        #split samples
                        sampleName = self.samples[i].split('_')[0]
                        sampleNumber = self.samples[i].split('_')[1]
                        sampleMeasure = self.samples[i].split('_')[2]
                        print(self.filePath + '/' + sampleName + '-' + str(h) + 'min-' + str(sampleNumber) + '-' + str(sampleMeasure) + '.dat')
                        file_path = self.filePath + '/' + sampleName + '-' + str(h) + 'min-' + str(sampleNumber) + '-' + str(sampleMeasure) + '.dat'
                        if os.path.exists(file_path):
                            iv = Data.Data.getDataListPsc(file_path)
                            self.ax.plot(iv[1], iv[0], label=str(h) + 'min-' + str(sampleNumber) + '-' + str(sampleMeasure))

                            self.ax.set_xlabel('Voltage [V]')
                            self.ax.set_ylabel('Current [A]')
                            self.ax.set_title('Light IV')
                            # Put a legend to the right of the current axis
                            self.ax.legend(loc='center left', bbox_to_anchor=(1, 0.5))
        else:
            for i in range(0, len(self.samples)):
                if self.sampleCb[i].get() == 1:
                    for h in self.hours:
                        if os.path.exists(self.filePath + '/' + self.samples[i] + '/' + str(h) + '.csv'):
                            iv = Data.Data.getDataListSm(self.filePath + '/' + self.samples[i] + '/' + str(h) + '.csv')
                            self.bx.plot(iv[1], iv[0], label=str(h) + 'min')
                            self.bx.set_xlabel('Voltage [V]')
                            self.bx.set_ylabel('Current [A]')
                            self.bx.set_title('Dark IV')
                            # Put a legend to the right of the current axis
                            self.bx.legend(loc='center left', bbox_to_anchor=(1, 0.5))

                    if os.path.exists(self.filePath + '/' + self.samples[i] + '/' + 'Rsh.csv'):
                        self.rsh = Data.Data.getDataListSm(self.filePath + '/' + self.samples[i] + '/' + 'Rsh.csv')
                        self.ax.plot(self.rsh[1], self.rsh[0], label=str(self.samples[i]))
                        self.ax.set_xlabel('Time [min]')
                        self.ax.set_ylabel('Rsh [Ohm]')
                        self.ax.set_title('Rsh')
                        # Put a legend to the right of the current axis
                        self.ax.legend(loc='center left', bbox_to_anchor=(1, 0.5))

        self.canvas.draw()
        self.canvas2.draw()

        self.table()

    def table(self):
        for widget in self.tFrame.winfo_children():
            widget.destroy()
        if self.machineRb.get() == 1:
            for i in range(0, len(self.samples)):
                if self.sampleCb[i].get() == 1:
                    self.tFrameSample = ttk.Frame(self.tFrame, style='My.TFrame')
                    self.tFrameSample.pack()

                    # dirname = self.filePath
                    # filespec = str(self.subfolders[-1]) + '-' + '*' + self.wbName
                    # file = glob.glob(os.path.join(dirname,filespec))[0]
                    if os.path.exists(self.newFile):
                        self.titleS = ttk.Label(self.tFrameSample, text=self.samples[i], style='My.TLabel').grid(sticky='NW',row=0,column=0,padx=5)
                        row = 2
                        columnH = 0
                        columnP = 2
                        columnI = 4
                        columnV = 6
                        columnF = 8

                        self.titleH = ttk.Label(self.tFrameSample, text='Time [h]', style='My.TLabel').grid(sticky='NW', row=row, column=columnH, padx=5)
                        ttk.Separator(self.tFrameSample, orient=VERTICAL, style='My.TSeparator').grid(row=row, column=1, rowspan=len(self.hours)+2, sticky='NS')
                        self.titleP = ttk.Label(self.tFrameSample, text='%PID [%]', style='My.TLabel').grid(sticky='NW', row=row, column=columnP, padx=5)
                        ttk.Separator(self.tFrameSample, orient=VERTICAL, style='My.TSeparator').grid(row=row, column=3,rowspan=len(self.hours) + 2, sticky='NS')
                        self.titleI = ttk.Label(self.tFrameSample, text='Isc [mA]', style='My.TLabel').grid(sticky='NW', row=row, column=columnI,padx=5)
                        ttk.Separator(self.tFrameSample, orient=VERTICAL, style='My.TSeparator').grid(row=row, column=5,rowspan=len(self.hours) + 2, sticky='NS')
                        self.titleV = ttk.Label(self.tFrameSample, text='Voc [mV]', style='My.TLabel').grid(sticky='NW', row=row, column=columnV,padx=5)
                        ttk.Separator(self.tFrameSample, orient=VERTICAL, style='My.TSeparator').grid(row=row, column=7,rowspan=len(self.hours) + 2,sticky='NS')
                        self.titleF = ttk.Label(self.tFrameSample, text='FF [%]', style='My.TLabel').grid(sticky='NW', row=row, column=columnF,padx=5)

                        l=0
                        for h in range(0, len(self.hours)):
                            # get %PID from Excel
                            wb = openpyxl.load_workbook(self.newFile, data_only=True)
                            activeSheet = wb[self.samples[i]]
                            if activeSheet.cell(row=h+2, column=17).value is not None:
                                labelP = round(activeSheet.cell(row=h+2, column=17).value,2)
                            else:
                                labelP = '' # label over label -> update
                            if activeSheet.cell(row=h + 2, column=5).value is not None:
                                labelI = activeSheet.cell(row=h + 2, column=5).value
                            else:
                                labelI = ''
                            if activeSheet.cell(row=h + 2, column=7).value is not None:
                                labelV = activeSheet.cell(row=h + 2, column=7).value
                            else:
                                labelV = ''
                            if activeSheet.cell(row=h + 2, column=8).value is not None:
                                labelF = activeSheet.cell(row=h + 2, column=8).value
                            else:
                                labelF = ''
                            l = h + 2+row
                            ttk.Label(self.tFrameSample, text=str(self.hours[h]), style='My.TLabel').grid(row=l, column=columnH, padx=5, sticky='WE')
                            ttk.Label(self.tFrameSample, text=labelP, style='My.TLabel').grid(row=l, column=columnP, padx=5, sticky='WE')
                            ttk.Label(self.tFrameSample, text=labelI, style='My.TLabel').grid(row=l, column=columnI, padx=5, sticky='WE')
                            ttk.Label(self.tFrameSample, text=labelV, style='My.TLabel').grid(row=l, column=columnV, padx=5, sticky='WE')
                            ttk.Label(self.tFrameSample, text=labelF, style='My.TLabel').grid(row=l, column=columnF, padx=5, sticky='WE')

                        ttk.Separator(self.tFrameSample, orient=HORIZONTAL, style='My.TSeparator').grid(row=l+1,columnspan=9,sticky='WE',padx=5)

        elif self.machineRb.get() == 2:
            print('thin film')
            for i in range(0, len(self.samples)):
                if self.sampleCb[i].get() == 1:
                    self.tFrameSample = ttk.Frame(self.tFrame, style='My.TFrame')
                    self.tFrameSample.pack()

                    if os.path.exists(self.newFile):
                        self.titleS = ttk.Label(self.tFrameSample, text=self.samples[i], style='My.TLabel').grid(sticky='NW',row=0,column=0,padx=5)
                        row = 2
                        columnH = 0
                        columnP = 2
                        columnV = 4
                        columnF = 6
                        columnE = 8

                        self.titleH = ttk.Label(self.tFrameSample, text='Time [h]', style='My.TLabel').grid(sticky='NW', row=row, column=columnH, padx=5)
                        ttk.Separator(self.tFrameSample, orient=VERTICAL, style='My.TSeparator').grid(row=row, column=1, rowspan=len(self.hours)+2, sticky='NS')
                        self.titleP = ttk.Label(self.tFrameSample, text='%PID [%]', style='My.TLabel').grid(sticky='NW', row=row, column=columnP, padx=5)
                        ttk.Separator(self.tFrameSample, orient=VERTICAL, style='My.TSeparator').grid(row=row, column=3,rowspan=len(self.hours) + 2, sticky='NS')
                        self.titleV = ttk.Label(self.tFrameSample, text='Voc [mV]', style='My.TLabel').grid(sticky='NW', row=row, column=columnV,padx=5)
                        ttk.Separator(self.tFrameSample, orient=VERTICAL, style='My.TSeparator').grid(row=row, column=5,rowspan=len(self.hours) + 2, sticky='NS')
                        self.titleF = ttk.Label(self.tFrameSample, text='FF [%]', style='My.TLabel').grid(sticky='NW', row=row, column=columnF,padx=5)
                        ttk.Separator(self.tFrameSample, orient=VERTICAL, style='My.TSeparator').grid(row=row, column=7,rowspan=len(self.hours) + 2,sticky='NS')
                        self.titleE = ttk.Label(self.tFrameSample, text='Eff [%]', style='My.TLabel').grid(sticky='NW', row=row, column=columnE,padx=5)

                        l=0
                        for h in range(0, len(self.hours)):
                            # get %PID from Excel
                            wb = openpyxl.load_workbook(self.newFile, data_only=True)
                            activeSheet = wb[self.samples[i]]
                            if activeSheet.cell(row=h+2, column=6).value is not None:
                                labelP = round(activeSheet.cell(row=h+2, column=6).value,2)
                            else:
                                labelP = '' # label over label -> update
                            if activeSheet.cell(row=h + 2, column=2).value is not None:
                                labelV = activeSheet.cell(row=h + 2, column=2).value
                            else:
                                labelV = ''
                            if activeSheet.cell(row=h + 2, column=4).value is not None:
                                labelF = activeSheet.cell(row=h + 2, column=4).value
                            else:
                                labelF = ''
                            if activeSheet.cell(row=h + 2, column=5).value is not None:
                                labelE = activeSheet.cell(row=h + 2, column=5).value
                            else:
                                labelE = ''
                            l = h + 2+row
                            ttk.Label(self.tFrameSample, text=str(self.hours[h]), style='My.TLabel').grid(row=l, column=columnH, padx=5, sticky='WE')
                            ttk.Label(self.tFrameSample, text=labelP, style='My.TLabel').grid(row=l, column=columnP, padx=5, sticky='WE')
                            ttk.Label(self.tFrameSample, text=labelV, style='My.TLabel').grid(row=l, column=columnV, padx=5, sticky='WE')
                            ttk.Label(self.tFrameSample, text=labelF, style='My.TLabel').grid(row=l, column=columnF, padx=5, sticky='WE')
                            ttk.Label(self.tFrameSample, text=labelE, style='My.TLabel').grid(row=l, column=columnE, padx=5, sticky='WE')

                        ttk.Separator(self.tFrameSample, orient=HORIZONTAL, style='My.TSeparator').grid(row=l+1,columnspan=9,sticky='WE',padx=5)
        else:
            ttk.Separator(self.tFrame, orient=HORIZONTAL, style='My.TSeparator').grid(row=1, columnspan=15, sticky='WE', padx=5)
            columnH = 0
            for i in range(0, len(self.hours)):
                ttk.Label(self.tFrame, text=str(self.hours[i]), style='My.TLabel').grid(row=i + 2, column=columnH,
                                                                                        padx=5, sticky='WE')

            self.titleH = ttk.Label(self.tFrame, text='Time [min]', style='My.TLabel').grid(sticky='NW', row=0,column=columnH, padx=5)
            ttk.Separator(self.tFrame, orient=VERTICAL, style='My.TSeparator').grid(row=0, column=1,rowspan=len(self.hours) + 2,sticky='NS')

            for s in range(0, len(self.samples)):
                if self.sampleCb[s].get() == 1:
                    if os.path.exists(self.filePath + '/' + self.samples[s] + '/' + 'Rsh.csv'):
                        self.rsh = Data.Data.getDataListSm(self.filePath + '/' + self.samples[s] + '/' + 'Rsh.csv')
                        columnR = 2+s*2
                        ttk.Separator(self.tFrame, orient=VERTICAL, style='My.TSeparator').grid(row=0, column=columnR+1, rowspan=len(self.hours)+2, sticky='NS')
                        self.titleR = ttk.Label(self.tFrame, text=self.samples[s], style='My.TLabel').grid(sticky='NW', row=0, column=columnR, padx=5)
                        i = 0
                        for r in range(0, len(self.rsh[1])):
                            if int(self.rsh[1][r]) == int(self.hours[i]):
                                labelR = self.rsh[0][r]
                                ttk.Label(self.tFrame, text=labelR, style='My.TLabel').grid(row=i+2, column=columnR, padx=5, sticky='WE')
                                i = i + 1
                                if i >= len(self.hours):
                                    break

    def selectMachine(self):
        if self.machineRb.get() == 1:
            self.iv.config(state=NORMAL)
            self.eqe.config(state=NORMAL)
            self.photo.config(state=NORMAL)
            self.everyTimeLabel.config(text='h')
            self.everyTimeLabel.update()
        else:
            self.iv.config(state=DISABLED)
            self.eqe.config(state=DISABLED)
            self.photo.config(state=DISABLED)
            self.everyTimeLabel.config(text='min')
            self.everyTimeLabel.update()

    def pickFile(self, event):
        """
        select a .xlsx file via the filedialog
        is called by fileBtn
        """
        caller = event.widget
        print(caller)
        self.cbFrame.destroy()
        self.cbFrame = None
        self.cbFrame = ttk.Frame(self.tabInfo, style='My.TFrame')
        self.cbFrame.grid(row=0, columnspan=5, sticky='W', padx=5)
        self.filePath = ''
        self.samples = []
        self.wbName = ''
        self.hours = []
        self.sampleRb = tk.IntVar()

        self.filename = tk.filedialog.askopenfilename(initialdir="/", title="Select file",filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))
        if self.filename != None:
                self.fileVar.set(self.filename)
                self.wbName = self.filename.split('/')[-1]
                self.filePath = self.filename.replace(self.wbName, '')

    def getSamples(self):
        self.samples=[]
        self.sampleNames=[]
        self.sampleNumbers=[]
        self.sampleMeasures=[]
        #LOANA
        if self.machineRb.get() == 1:
            for f in listdir(self.filePath):
                if os.path.isdir(self.filePath + f):
                    for g in listdir(self.filePath + f):
                        if os.path.isdir(self.filePath + f + '/' + g):
                            self.samples.append(g)
                break
        #PME
        elif self.machineRb.get() == 2:
            files = [f.name for f in os.scandir(self.filePath)]
            fileNames = []
            for f in files:
                if f.endswith('.dat'):
                    fileNames.append(f[:-4])
            fileNamesSorted = WorkbookLayout.WorkbookLayout.natural_sort(fileNames)
            for f in fileNamesSorted:
                fileNamesSplitted = f.split('-')
                if fileNamesSplitted[0] not in self.sampleNames:
                    self.sampleNames.append(fileNamesSplitted[0])
                if int(fileNamesSplitted[1][:-3]) not in self.hours:
                    self.hours.append(int(fileNamesSplitted[1][:-3]))
                if len(fileNamesSplitted) == 4:
                    if int(fileNamesSplitted[2]) not in self.sampleNumbers:
                        self.sampleNumbers.append(int(fileNamesSplitted[2]))
                    if fileNamesSplitted[3] not in self.sampleMeasures:
                        self.sampleMeasures.append(fileNamesSplitted[3])
                else:
                    if int(fileNamesSplitted[3]) not in self.sampleNumbers:
                        self.sampleNumbers.append(int(fileNamesSplitted[3]))
                    if fileNamesSplitted[4] not in self.sampleMeasures:
                        self.sampleMeasures.append(fileNamesSplitted[4])
            self.sampleNames = list(set(self.sampleNames))
            self.hours = list(set(self.hours))
            self.sampleNumbers = list(set(self.sampleNumbers))
            self.sampleMeasures = list(set(self.sampleMeasures))
            self.sampleMeasures.sort(key=int)
            self.hours.sort(key=int)
            for n in self.sampleNames:
                for i in self.sampleNumbers:
                    for m in self.sampleMeasures:
                        self.samples.append(n + '_' + str(i) + '_' + m)
        #K2400
        else:
            for f in listdir(self.filePath):
                if os.path.isdir(self.filePath + f):
                    self.samples.append(f)
        #Checkbuttons
        if self.samples != []:
            row=0
            column=0
            for i in range(0, len(self.samples)):
                cbValue = tk.IntVar()
                self.sampleCb.append(cbValue)
                cb = ttk.Checkbutton(self.cbFrame, text=self.samples[i], variable=self.sampleCb[i], onvalue=1, offvalue=0, style='My.TCheckbutton')
                #cb.pack(side='left', fill=None, expand=False, padx=(0,5))
                cb.grid(row=row, column=column, sticky='W', padx=5)
                column = column+1
                if column > 7:
                    column = 0
                    row = row+1

    def begin(self):
        rb = self.machineRb.get()
        self.getSamples()
        if rb == 1:
            self.beginBifi()
        elif rb == 2:
            self.beginPsc()
        elif rb == 3:
            self.beginSm()
        else:
            self.errorLabel.config(text='Error with machine selection')

    def beginBifi(self):
        """
        check if all variables are set correctly and starts the beginBifi function of the Main class
        is called by beginBtn
        """
        if self.hourRb.get() == 2 and(self.hourInput.get() == '' or self.hourInput2.get() == ''):
            self.errorLabel.config(text='Error : No hour')
        elif self.hourRb.get() == 2 and (not self.hourInput.get().isdigit() or not self.hourInput2.get().isdigit()):
            self.errorLabel.config(text='Error : Hour must be a number')
        elif self.hourRb.get() == 2 and (int(self.hourInput.get()) >= int(self.hourInput2.get())):
            self.errorLabel.config(text='Error : Hour interval is not correct')
        else:
            # if begin function is done : show 'file saved' and save the configuration if not already in config.txt
            if(self.fileInput.get().endswith('.xlsx')):
                self.errorLabel.config(text='Running...')
                self.errorLabel.update()

                #calculate hours
                #all hours
                self.hours = AlgemeneInfo.AlgemeneInfo.calculateHours(self.wbName, self.filePath)
                self.subfolders = [f.name for f in os.scandir(self.filePath) if f.is_dir()]
                self.subfolders.sort(key=float)
                #between hours
                if self.hourRb.get() != 1:
                    times = self.hours
                    self.hours = []
                    allsubfolders = self.subfolders
                    self.subfolders = []
                    for h in range(0, len(times)):
                        if int(self.hourInput.get()) <= int(times[h]) and int(times[h]) <= int(self.hourInput2.get()):
                            self.hours.append(int(times[h]))
                            self.subfolders.append(allsubfolders[h])
                #data every
                if self.timeCb.get() == 1:
                    threshold = int(self.timeInt.get())
                    times = self.hours
                    self.hours = [times[0]]
                    allsubfolders = self.subfolders
                    self.subfolders = [allsubfolders[0]]
                    lastVal = 0
                    for h in range(0, len(times)):
                        if int(times[h]) >= int(lastVal) + int(threshold):
                            self.hours.append(int(times[h]))
                            self.subfolders.append(allsubfolders[h])
                            lastVal = int(times[h])

                dataType = ''
                if (self.ivCb.get()):
                    dataType = dataType + 'iv-'
                if (self.eqeCb.get()):
                    dataType = dataType + 'eqe-'
                self.newFile = self.filePath + str(self.subfolders[0]) + str(self.subfolders[-1]) + '-' + dataType + self.wbName
                if Main.Main.beginBifi(self, self.hours, self.subfolders, self.wbName, self.filePath, self.ivCb.get(), self.eqeCb.get(), self.photoCb.get(), self.newFile):
                    self.errorLabel.config(text='File saved')

                else:
                    self.errorLabel.config(text='Error: Something went wrong')
            else:
                self.errorLabel.config(text='Error : Not a .xlsx file')

    def beginPsc(self):
        """
        check if all variables are set correctly and starts the beginPsc function of the Main class
        is called by beginBtnPsc
        """
        if self.hourRb.get() == 2 and(self.hourInput.get() == '' or self.hourInput2.get() == ''):
            self.errorLabel.config(text='Error : No hour')
        elif self.hourRb.get() == 2 and (not self.hourInput.get().isdigit() or not self.hourInput2.get().isdigit()):
            self.errorLabel.config(text='Error : Hour must be a number')
        elif self.hourRb.get() == 2 and (int(self.hourInput.get()) >= int(self.hourInput2.get())):
            self.errorLabel.config(text='Error : Hour interval is not correct')
        else:
            if(self.fileInput.get().endswith('.xlsx')):
                self.errorLabel.config(text='Running...')
                self.errorLabel.update()

                if self.hourRb.get() == 2:
                    times = self.hours
                    self.hours = []
                    for h in times:
                        if int(self.hourInput.get()) <= int(h) and int(h) <= int(self.hourInput2.get()):
                            self.hours.append(int(h))
                if self.timeCb.get() == 1:
                    threshold = int(self.timeInt.get())
                    times = self.hours
                    self.hours = [times[0]]
                    lastVal = times[0]
                    for h in times:
                        if int(h) >= int(lastVal) + int(threshold):
                            self.hours.append(int(h))
                            lastVal = int(h)
                self.newFile = self.filePath + str(self.hours[0]) + '-' + str(self.hours[-1]) + '-' + self.wbName
                if Main.Main.beginPsc(self, self.wbName, self.filePath, self.hours, self.samples, self.newFile):
                    self.errorLabel.config(text='File saved')
                else:
                    self.errorLabel.config(text='Error: Something went wrong')
            else:
                self.errorLabel.config(text='Error : Not a .xlsx file')

    def beginSm(self):
        """
        check if all variables are set correctly and starts the beginSm function of the Main class
        is called by beginBtnSm
        """
        modulePath = self.filePath + self.samples[0]
        files = [f.name for f in os.scandir(modulePath)]
        for f in files:
            if f.endswith('.csv') and not f.startswith('Rsh'):
                self.hours.append(f[:-4])
        self.hours.sort(key=int)

        if self.hourRb.get() == 2 and(self.hourInput.get() == '' or self.hourInput2.get() == ''):
            self.errorLabel.config(text='Error : No hour')
        elif self.hourRb.get() == 2 and (not self.hourInput.get().isdigit() or not self.hourInput2.get().isdigit()):
            self.errorLabel.config(text='Error : Hour must be a number')
        elif self.hourRb.get() == 2 and (int(self.hourInput.get()) >= int(self.hourInput2.get())):
            self.errorLabel.config(text='Error : Hour interval is not correct')
        else:
            if self.hourRb.get() == 2:
                times = self.hours
                self.hours = []
                for h in times:
                    if int(self.hourInput.get()) <= int(h) and int(h) <= int(self.hourInput2.get()):
                        self.hours.append(int(h))
            if self.timeCb.get() == 1:
                threshold = int(self.timeInt.get())
                times = self.hours
                self.hours = [times[0]]
                lastVal = times[0]
                for h in times:
                    if int(h) >= int(lastVal) + int(threshold):
                        self.hours.append(int(h))
                        lastVal = int(h)
            self.newFile = self.filePath + str(self.hours[0]) + '-'+ str(self.hours[-1]) + '-' + self.wbName

            if(self.fileInput.get().endswith('.xlsx')):
                self.errorLabel.config(text='Running...')
                self.errorLabel.update()
                if Main.Main.beginSm(self, self.wbName, self.filePath, self.hours, self.newFile):
                    self.errorLabel.config(text='File saved')
                else:
                    self.errorLabel.config(text='Error: Something went wrong')
            else:
                self.errorLabel.config(text='Error : Not a .xlsx file')



root = Tk()
root.title("PID data automation")
root.tk.call('wm', 'iconphoto', root._w, PhotoImage(file='Monocrystalline-solar-cell.png'))
app = Application(master=root)
app.mainloop()