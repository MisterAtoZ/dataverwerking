import openpyxl

class AlgemeneInfo():
    def datasheet(workbook, hours, sheetNames, dataEx, subfolders):
        """
        calculates the hours and set these values in the general sheet,
        puts the data from dataEx in the correct sheets,
        and looks which hours are already filled in and thus which hours still need to be filled in
        :param workbook: Excel workbook
        :param sheetNames: list of sheet names
        :param dataEx: data exchange file
        :param maxHour: value of the maximum hour that needs to be filled in
        :param subfolders: list of subfolders
        :return: list containing the begin row and list of hours
        """
        wb = workbook
        data = dataEx['data-exchange']
        # hours = AlgemeneInfo.calculateHours(wb)
        print('algemene info')

        # fill in dataEx data in correct sheets
        for n in range(0, len(sheetNames), 1):
            print(sheetNames[n])
            sheet = wb[sheetNames[n]]
            nextRow = 1
            # calculate the first empty row
            for row in range(1, sheet.max_row+2, 1):
                if sheet.cell(row=row, column=1).value == None:
                    nextRow = row
                    break
            begin = nextRow
            # fill in the data from dataEx
            for rows in range(begin, 2+len(hours), 1):
                hourInt = int(round(hours[rows-2],0))
                hourRounded = round(hours[rows-2],2)
                # if hourInt <= int(maxHour) or maxHour == -1:
                print(hourRounded)
                sheet.cell(row=nextRow, column=1).value = hourRounded
                for i in range(4,data.max_row,len(sheetNames)):
                    if (str(data.cell(row=i, column=1).value) == str(hourInt)) or (str(data.cell(row=i, column=1).value) == str(subfolders[rows-2])):
                        for j in range(i,i+len(sheetNames),1):
                            if (str(data.cell(row=j, column=2).value) == sheetNames[n]):
                                for k in range(2, 13, 1):
                                    sheet.cell(row=nextRow, column=k).value = data.cell(row=j, column=k + 7).value
                                if (sheet.cell(row=nextRow,column=4).value is not None) and (sheet.cell(row=2, column=4).value is not None):
                                    sheet.cell(row=nextRow, column=17).value = 100 - 100 * sheet.cell(row=nextRow,column=4).value / sheet.cell(row=2, column=4).value
                                nextRow = nextRow + 1
                # else:
                #     hours = hours[:rows-2]
                #     break

        for i in range(0, len(hours),1):
            hours[i] = int(round(hours[i],0))
        return [begin, hours]

    def calculateHours(wbName, path):

        wb = openpyxl.load_workbook(path + wbName, data_only=True)
        generalSheet = wb['General']

        # calculate the row with the first stressed hour
        for i in range(1, generalSheet.max_row, 1):
            if generalSheet.cell(row=i, column=5).value == 'Acc. Hours  [h]':
                beginRow = i + 2
                print('beginRow : ' + str(beginRow))

        # calculate hours
        previousHour = 0
        hours = [0]
        # fill in the hours in general sheet and make a list of all these hours
        # first check if a date is filled in.
        #   No : break out of for loop because no hours come after
        #   Yes : check if an hour is filled in. Yes : don't have to calculate hour but still append to list
        #                                        No : calculate hour and append to list
        for i in range(beginRow, generalSheet.max_row, 1):
            if generalSheet.cell(row=i, column=1).value == None:
                break
            if generalSheet.cell(row=i, column=5).value != None:
                previousHour = generalSheet.cell(row=i, column=5).value
            else:
                date1 = generalSheet.cell(row=i, column=1).value
                date2 = generalSheet.cell(row=i - 1, column=1).value
                timeOut = generalSheet.cell(row=i, column=2).value
                timeIn = generalSheet.cell(row=i - 1, column=3).value

                date1 = date1.replace(hour=timeOut.hour, minute=timeOut.minute)
                date2 = date2.replace(hour=timeIn.hour, minute=timeIn.minute)

                seconds = (date1 - date2).total_seconds()
                interval = seconds / 3600

                generalSheet.cell(row=i, column=4).value = interval
                generalSheet.cell(row=i, column=5).value = interval + previousHour

                previousHour = generalSheet.cell(row=i, column=5).value
            hours.append(previousHour)
        print (hours)

        wb.save(path + wbName)
        return hours


    def datasheetPsc(activeSheet, times, file_path_ini, t):
        data = ['Voc', 'Jsc', 'FF', 'Eff']
        i = 0
        activeSheet.cell(row=2+t, column=1).value = str(times[t]) + ' min'
        with open(file_path_ini, 'r') as file:
           for line in file:
               if i < len(data):
                   if line.startswith(data[i]):
                       split = line.split('"')
                       activeSheet.cell(row=2+t, column=2+i).value = split[1]
                       i = i + 1
               else:
                   #calculate PID
                   activeSheet.cell(row=2+t, column=6).value = 100-100*float(activeSheet.cell(row=2+t, column=4).value)/float(activeSheet.cell(row=2,column=4).value)
                   break