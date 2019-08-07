import AlgemeneInfo
import Data
import Grafieken
import WorkbookLayout
import Photo
import openpyxl
import os
from os import listdir

class Main():
    def beginBifi(self, hours, subfolders, wbName, path, ivCb, eqeCb, photoCb, newFile):
        try:
            graphNames = ['%PID', 'FF', 'Voc', 'Isc']
            sheetNames = [f.name for f in os.scandir(path + subfolders[0]) if f.is_dir()]
            wb = openpyxl.load_workbook(path + wbName, data_only=True)
            begin = 0

            if os.path.exists(path):
                for f in listdir(path):
                    if f.startswith('data-exchange') and f.endswith('.xlsx'):
                        data_exchange = f
                        dataEx = openpyxl.load_workbook(path + data_exchange)
                        WorkbookLayout.WorkbookLayout.makeSheets(wb, dataEx, graphNames, sheetNames, eqeCb)
                        info = AlgemeneInfo.AlgemeneInfo.datasheet(wb, hours, sheetNames, dataEx, subfolders)
                        begin = info[0]
                        hours = info[1]
            if begin == 0 and os.path.exists(path + subfolders[len(subfolders) - 1]):
                for f in listdir(path + subfolders[len(subfolders) - 1]):
                    if f.startswith('data-exchange') and f.endswith('.xlsx'):
                        data_exchange = f
                        dataEx = openpyxl.load_workbook(path + subfolders[len(subfolders) - 1] + '/' + data_exchange)
                        WorkbookLayout.WorkbookLayout.makeSheets(wb, dataEx, graphNames, sheetNames, eqeCb)
                        info = AlgemeneInfo.AlgemeneInfo.datasheet(wb, hours, sheetNames, dataEx, subfolders)
                        begin = info[0]
                        hours = info[1]
            if begin == 0:
                print('Error : data_exchange file does not exist')
                return False

            if(photoCb):
                wbNameImg = wbName.replace('.xlsx','')
                Photo.Photo.makeTitle(wbNameImg)

            for n in range(0, len(sheetNames), 1):
                if(photoCb):
                    Photo.Photo.makeTitle(sheetNames[n])
                # begin = eerste rij die ingevult moet worden (dus begin-1), maar de hours beginnen pas op rij 2 (dus nog eens -1) => begin-2
                ivHours = []
                eqeHours = []
                for hour in range(begin-2, len(hours), 1):
                    newPath = path + str(subfolders[hour]) + '/' + sheetNames[n]

                    if(photoCb):
                        photoPath = path + str(subfolders[hour]) + '/'
                        imgHour = str(hours[hour]) + ' h'
                        imgSampleResised = photoPath + sheetNames[n] + '_resised'
                        for f in listdir(photoPath):
                            if (f.endswith(str(sheetNames[n]) + '.JPG') or f.endswith(str(sheetNames[n]) + '.jpg')):
                                photo = f
                                Photo.Photo.resize(photoPath + photo, imgSampleResised)
                        #HORIZONTALE TOEVOEGING
                        if n == 0:
                            Photo.Photo.makeTitle(imgHour)
                            Photo.Photo.merge_image(str(wbNameImg) + '.jpg', imgHour + '.jpg', False, str(wbNameImg))
                            os.remove(imgHour + '.jpg')
                        Photo.Photo.merge_image(str(sheetNames[n]) + '.jpg', imgSampleResised + '.jpg', False, str(sheetNames[n]))
                        #remove files
                        os.remove(imgSampleResised + '.jpg')


                    activeSheet = wb[sheetNames[n]]

                    ivPath = newPath + '/IV/'
                    eqePath = newPath + '/IQE/'

                    drk = []
                    lgt = []

                    if(ivCb):
                        #data uit file halen
                        if os.path.exists(ivPath):
                            if os.path.exists(ivPath + sheetNames[n] + '.drk'):
                                drk = Data.Data.getDataList(str(ivPath) + sheetNames[n] + '.drk')
                            else:
                                print(str(ivPath) + '.drk file bestaat niet')
                            if os.path.exists(ivPath + sheetNames[n] + '.lgt'):
                                lgt = Data.Data.getDataList(str(ivPath) + sheetNames[n] + '.lgt')
                            else:
                                print(str(ivPath) + sheetNames[n] + ' .lgt file bestaat niet')
                        else:
                            print(str(ivPath) + ' bestaat niet')

                        if drk != [] and lgt != []:
                            WorkbookLayout.WorkbookLayout.setIV(activeSheet, hours[hour], drk, lgt)
                            ivHours.append(hours[hour])


                    # ------ EQE ------
                    if(eqeCb):
                        eqeSheet = 'EQE_' + sheetNames[n]
                        activeSheet = wb[eqeSheet]

                        #data uit file halen
                        eqeFile = ''
                        if os.path.exists(eqePath):
                            for f in listdir(eqePath):
                                if f.startswith(sheetNames[n]) and f.endswith('.eqe'):
                                    eqeFile = f
                            if eqeFile != '':
                                eqe = Data.Data.getDataList(str(eqePath) + eqeFile)[0]
                                WorkbookLayout.WorkbookLayout.setEQE(activeSheet, hours[hour], eqe)
                                eqeHours.append(hours[hour])
                            else:
                                print(str(eqePath) + ' file bestaat niet')
                        else:
                            print(str(eqePath) + ' bestaat niet')

                if(photoCb):
                    #VERTICALE TOEVOEGING
                    Photo.Photo.merge_image(str(wbNameImg) + '.jpg', str(sheetNames[n]) + '.jpg', True, str(wbNameImg))
                    os.remove(str(sheetNames[n]) + '.jpg')

                #grafieken
                if(ivCb):
                    Grafieken.Grafieken.makeChart(sheetNames[n], wb, ivHours)
                if(eqeCb):
                    Grafieken.Grafieken.makeChart('EQE_' + sheetNames[n], wb, eqeHours)

            Grafieken.Grafieken.makeSeperateGraphs(wb, graphNames, sheetNames, hours)

            if (photoCb):
                # os.rename(os.path.dirname(os.path.dirname(path)) + '/' + str(wbNameImg) + '.jpg', path + str(hours[-1]) + '-' + str(wbNameImg) + '.jpg')
                if os.path.exists(path + str(subfolders[-1]) + '-' + str(wbNameImg) + '.jpg'):
                    os.remove(path + str(subfolders[-1]) + '-' + str(wbNameImg) + '.jpg')
                os.rename(str(wbNameImg) + '.jpg', path + str(subfolders[0]) + str(subfolders[-1]) + '-' + str(wbNameImg) + '.jpg')

            wb.save(newFile)
            print('Saving...')
            return True

        except:
            # remove all photo's which are already made before the error
            if(photoCb):
                for f in listdir(os.path.dirname(os.path.dirname(path))):
                    if f.endswith('.jpg'):
                        os.remove(f)
            pass
        return False

    def beginPsc(self, wbName, filePath, times, sheetNames, newFile):
        # try:
        wb = openpyxl.load_workbook(filePath + wbName, data_only=True)
        graphNames = ['%PID']
        WorkbookLayout.WorkbookLayout.makeSheetsPsc(wb, graphNames, sheetNames)
        print(sheetNames)
        print(len(sheetNames))
        print(times)
        for sn in sheetNames:
            activeSheet = wb[sn]
            for t in range(0, len(times)):
                # pad+n-tmin-i-m.dat
                n = sn.split('_')[0]
                i = sn.split('_')[1]
                m = sn.split('_')[2]
                file_path_ini = filePath + n + '-' + str(times[t]) + 'min-' + str(i) + '-' + str(m) + '.ini'
                if os.path.exists(file_path_ini):
                    AlgemeneInfo.AlgemeneInfo.datasheetPsc(activeSheet, times, file_path_ini, t)
                file_path = filePath + n + '-' + str(times[t]) + 'min-' + str(i) + '-' + str(m) + '.dat'
                if os.path.exists(file_path):
                    iv = Data.Data.getDataListPsc(file_path)
                    WorkbookLayout.WorkbookLayout.setIVPsc(activeSheet, times[t], iv)
            print(sn)
            print(activeSheet.max_row)
            print(activeSheet.max_column)
            if (activeSheet.max_column == 6 and activeSheet.max_row == 1):
                wb.remove(activeSheet)
                print('sheet removed')
                continue
            Grafieken.Grafieken.makeChartPscSm(sn, wb, times, 'Psc')
        Grafieken.Grafieken.makeSeperateGraphsPsc(wb, sheetNames, times)
        print('Saving...')
        wb.save(newFile)

        return True

        # except:
        #     print('Error')
        #     pass
        #
        # return False

    def beginSm(self, wbName, filePath, hours, newFile):
        try:
            wb = openpyxl.load_workbook(filePath + wbName, data_only=True)
            sheetNames = [f.name for f in os.scandir(filePath) if f.is_dir()]
            WorkbookLayout.WorkbookLayout.makeSheetsSm(wb, sheetNames)
            for n in sheetNames:
                activeSheet = wb[n]
                modulePath = filePath + '/' + n
                rsh = Data.Data.getDataListSm(str(modulePath) + '/' + 'Rsh.csv')
                WorkbookLayout.WorkbookLayout.setRsh(wb, rsh, n)
                for h in hours:
                    if os.path.exists(str(modulePath) + '/' + str(h) + '.csv'):
                        iv = Data.Data.getDataListSm(str(modulePath) + '/' + str(h) + '.csv')
                        WorkbookLayout.WorkbookLayout.setIVSm(activeSheet, str(h), iv)
                Grafieken.Grafieken.makeChartPscSm(n, wb, hours, 'Sm')
            Grafieken.Grafieken.makeChartRsh(wb, sheetNames)

            wb.save(newFile)
            print('Saving...')
            return True

        except:
            print('Error')
            pass
        return False