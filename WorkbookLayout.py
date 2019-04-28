import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment

import re


class WorkbookLayout():

    def makeSheets(workbook, dataEx, graphNames, sheetNames, eqeCb):
        """
        create the sheets in the Excel workbook
        :param workbook: Excel workbook
        :param dataEx: data exchange file used to set the text in row 1 which tells the information dataEx contains
        :param graphNames: list of sheets with only a graph
        :param sheetNames: list of the other sheet names, these are the sample names. EQE sheets of the samples are also created
        """
        wb = workbook
        data = dataEx['data-exchange']
        # fonts
        bold = Font(bold=True)

        # create the graph sheets
        for j in range(0,len(graphNames),1):
            if graphNames[j] not in wb.sheetnames:
                wb.create_sheet(graphNames[j])

        # create the sample sheets
        for n in range(0, len(sheetNames), 1):
            if sheetNames[n] not in wb.sheetnames:
                wb.create_sheet(sheetNames[n])
                sheet = wb[sheetNames[n]]
                sheet.cell(row=1, column=1).value = 'Time [h]'
                sheet.cell(row=1, column=1).font = bold
                # put the data from dataEx in the sheet
                for j in range(2, 13, 1):
                    sheet.cell(row=1, column=j).value = data.cell(row=2, column=j + 7).value + ' [' + data.cell(row=3, column=j + 7).value + ']'
                    sheet.cell(row=1, column=j).font = bold
                sheet.cell(row=1, column=17).value = '%PID [%]'
                sheet.cell(row=1, column=17).font = bold
            if(eqeCb):
                if 'EQE_' + sheetNames[n] not in wb.sheetnames:
                    wb.create_sheet('EQE_' + sheetNames[n])
                    sheet = wb['EQE_' + sheetNames[n]]
                    sheet.cell(row=2, column=1).value = 'wavelength (Lambda)'
                    for j in range(0,93,1):
                        sheet.cell(row=j+3, column=1).value = 280 + j*10


    def makeSheetsPsc(workbook, graphNames, sheetNames):
        """
        create the sheets in the Excel workbook
        :param workbook: Excel workbook
        :param graphNames: list of sheets with only a graph
        :param sheetNames: list of the other sheet names, these are the sample names
        """
        wb = workbook
        # create the graph sheets
        for j in range(0, len(graphNames), 1):
            if graphNames[j] not in wb.sheetnames:
                wb.create_sheet(graphNames[j])
        # create the sample sheets
        for n in range(0, len(sheetNames), 1):
            if sheetNames[n] not in wb.sheetnames:
                wb.create_sheet(sheetNames[n])

    def makeSheetsSm(workbook, sheetNames):
        """
        create the sheets in the Excel workbook
        :param workbook: Excel workbook
        :param sheetNames: list of the other sheet names, these are the sample names
        """
        wb = workbook
        # create the module sheets
        for n in range(0, len(sheetNames), 1):
            if sheetNames[n] not in wb.sheetnames:
                wb.create_sheet(sheetNames[n])

    def setIV(activeSheet, hour, drk, lgt):
        """
        set the IV values of 1 time frame in 1 sheet
        :param activesheet: Excel sheet
        :param hour: integer of the hour value which is printed above the data
        :param drk: list of the IV data of the dark measurements
        :param lgt: list of the IV data of the light measurements
        """

        # fonts
        bold = Font(bold=True)
        big = Font(size=14)

        vDark = drk[1]
        iDark = drk[0]
        vLight = lgt[1]
        iLight = lgt[0]

        # set hour text
        column1 = activeSheet.max_column + 2
        column2 = activeSheet.max_column + 5
        activeSheet.merge_cells(start_row=1, start_column=column1, end_row=1, end_column=column2)
        activeSheet.cell(row=1, column=column1).value = str(hour) + ' h'
        activeSheet.cell(row=1, column=column1).font = bold
        activeSheet.cell(row=1, column=column1).alignment = Alignment(horizontal='center')

        # set light dark text
        column3 = column1 + 1
        activeSheet.merge_cells(start_row=2, start_column=column1, end_row=2, end_column=column3)
        activeSheet.cell(row=2, column=column1).value = 'light'
        activeSheet.cell(row=2, column=column1).font = big
        activeSheet.cell(row=2, column=column1).alignment = Alignment(horizontal='center')

        column4 = column1 + 2
        column5 = column4 + 1
        activeSheet.merge_cells(start_row=2, start_column=column4, end_row=2, end_column=column5)
        activeSheet.cell(row=2, column=column4).value = 'dark'
        activeSheet.cell(row=2, column=column4).font = big
        activeSheet.cell(row=2, column=column4).alignment = Alignment(horizontal='center')

        # set iv text
        activeSheet.cell(row=3, column=column1).value = 'V'
        activeSheet.cell(row=3, column=column3).value = 'I'
        activeSheet.cell(row=3, column=column4).value = 'V'
        activeSheet.cell(row=3, column=column5).value = 'I'

        # set data
        # --------------- verschil lengte light / dark ? ---------------
        for j in range(0, len(vLight), 1):
            activeSheet.cell(row=j + 4, column=column1).value = vLight[j]
            activeSheet.cell(row=j + 4, column=column3).value = iLight[j]
        for j in range(0, len(vDark), 1):
            activeSheet.cell(row=j + 4, column=column4).value = vDark[j]
            activeSheet.cell(row=j + 4, column=column5).value = iDark[j]


    def setEQE(activeSheet, hour, eqe):
        """
        set the EQE data of 1 time frame in 1 sheet
        :param activeSheet: Excel sheet
        :param hour: integer of the hour value which is printed above the data
        :param eqe: list of the EQE values
        """
        # set titles
        column1 = activeSheet.max_column + 2
        column2 = activeSheet.max_column + 3
        activeSheet.merge_cells(start_row=1, start_column=column1, end_row=1, end_column=column2)
        activeSheet.cell(row=1, column=column1).value = str(hour) + ' h'

        activeSheet.cell(row=2, column=column1).value = 'EQE'
        activeSheet.cell(row=2, column=column2).value = 'EQE norm.'

        # set data
        for j in range(0, len(eqe), 1):
            activeSheet.cell(row=j + 3, column=column1).value = eqe[j]
            activeSheet.cell(row=j + 3, column=column2).value = eqe[j] / activeSheet.cell(row=j + 3, column=3).value

    def setIVPsc(activeSheet, min, iv):
        """
        set the IV values of 1 time frame in 1 sheet
        :param activeSheet: Excel sheet
        :param min: integer of the minutes value which is printed above the data
        :param iv: list of the IV values
        """

        # fonts
        bold = Font(bold=True)

        i = iv[0]
        v = iv[1]

        # set minutes text
        if activeSheet.max_column == 1:
            column1 = activeSheet.max_column
            column2 = activeSheet.max_column + 1
        else:
            column1 = activeSheet.max_column + 2
            column2 = activeSheet.max_column + 3
        activeSheet.merge_cells(start_row=1, start_column=column1, end_row=1, end_column=column2)
        activeSheet.cell(row=1, column=column1).value = str(min) + ' min'
        activeSheet.cell(row=1, column=column1).font = bold
        activeSheet.cell(row=1, column=column1).alignment = Alignment(horizontal='center')

        # set iv text
        activeSheet.cell(row=2, column=column1).value = 'I'
        activeSheet.cell(row=2, column=column2).value = 'V'

        # set data
        for j in range(0, len(i), 1):
            activeSheet.cell(row=j + 3, column=column1).value = i[j]
            activeSheet.cell(row=j + 3, column=column2).value = v[j]


    def setIVSm(activeSheet, hour, iv):
        """
        set the IV values of 1 time frame in 1 sheet
        :param activeSheet: Excel sheet
        :param hour: integer of the hours value which is printed above the data
        :param iv: list of the IV values
        """
        # fonts
        bold = Font(bold=True)

        i = iv[0]
        v = iv[1]

        # set hours text
        if activeSheet.max_column == 1:
            column1 = activeSheet.max_column
            column2 = activeSheet.max_column + 1
        else:
            column1 = activeSheet.max_column + 2
            column2 = activeSheet.max_column + 3
        activeSheet.merge_cells(start_row=1, start_column=column1, end_row=1, end_column=column2)
        activeSheet.cell(row=1, column=column1).value = str(hour) + ' h'
        activeSheet.cell(row=1, column=column1).font = bold
        activeSheet.cell(row=1, column=column1).alignment = Alignment(horizontal='center')

        # set iv text
        activeSheet.cell(row=2, column=column1).value = 'I'
        activeSheet.cell(row=2, column=column2).value = 'V'

        # set data
        for j in range(0, len(i), 1):
            activeSheet.cell(row=j + 3, column=column1).value = i[j]
            activeSheet.cell(row=j + 3, column=column2).value = v[j]

    def natural_sort(l):
        """
        sorts a list in a natural way, this means lower number values come first in the list
        :param l: list which needs to be sorted
        :return: sorted list
        source: https://stackoverflow.com/questions/4836710/does-python-have-a-built-in-function-for-string-natural-sort
        """
        convert = lambda text: int(text) if text.isdigit() else text.lower()
        alphanum_key = lambda key: [convert(c) for c in re.split('([0-9]+)', key)]
        return sorted(l, key=alphanum_key)