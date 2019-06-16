import openpyxl

class Grafieken():
    def makeChart(sheetname, workbook, hours):
        """
        creates the IV and EQE graphs of a sample
        :param sheetname: name of the sheet
        :param workbook: Excel workbook
        :param hours: list of all the hours
        """
        wb = workbook
        sheet = wb[sheetname]

        # EQE graphs
        if sheetname.startswith('EQE'):
            for j in range(0,2,1):
                chartObj = openpyxl.chart.ScatterChart()
                chartObj.x_axis.title = 'Wavelenght [nm]'
                chartObj.legend.position = 'b'
                chartObj.y_axis.scaling.max = 1.2
                chartObj.y_axis.scaling.min = 0

                if j == 0:
                    chartObj.y_axis.title = 'EQE [-]'
                    location = 'C20'
                else:
                    chartObj.y_axis.title = 'Normalized EQE [-]'
                    location = 'M20'

                beginCol = 3
                for i in range(1, sheet.max_column, 1):
                        if sheet.cell(row=1, column=i).value == '0 h':
                            beginCol = i
                            break

                for i in range(0, len(hours), 1):
                    xvalues = openpyxl.chart.Reference(sheet, min_col=1, min_row=3, max_col=1, max_row=sheet.max_row)
                    yvalues = openpyxl.chart.Reference(sheet, min_col=beginCol+(i*3)+j, min_row=3, max_col=3+(i*3)+j, max_row=sheet.max_row)
                    seriesObj = openpyxl.chart.Series(yvalues, xvalues, title=str(hours[i]) + ' h')
                    chartObj.append(seriesObj)

                sheet.add_chart(chartObj, location)

        # IV graphs
        else:
            for j in range(0,2,1):
                chartObj = openpyxl.chart.ScatterChart()
                chartObj.x_axis.title = 'Voltage [V]'
                chartObj.y_axis.title = 'Current [A]'
                chartObj.x_axis.scaling.min = 0
                chartObj.legend.position = 'b'
                if j == 0:
                    chartObj.title = 'Light IV'
                    chartObj.y_axis.scaling.max = 0;
                    location = 'C20'
                else:
                    chartObj.title = 'Dark IV'
                    chartObj.y_axis.scaling.max = 10
                    location = 'J20'

                beginCol = 19
                for i in range(17,sheet.max_column,1):
                    if sheet.cell(row=1,column=i).value == '0 h':
                        beginCol = i
                        break

                for i in range(0, len(hours), 1):
                    xvalues = openpyxl.chart.Reference(sheet,min_col=beginCol+(i*5)+(j*2),min_row=4,max_col=beginCol+(i*5)+(j*2),max_row=sheet.max_row)
                    yvalues = openpyxl.chart.Reference(sheet,min_col=beginCol+1+(i*5)+(j*2),min_row=4,max_col=beginCol+1+(i*5)+(j*2),max_row=sheet.max_row)
                    seriesObj = openpyxl.chart.Series(yvalues, xvalues, title=str(hours[i]) + ' h')
                    chartObj.append(seriesObj)

                sheet.add_chart(chartObj, location)

    def makeSeperateGraphs(workbook, graphNames, sheetNames, hours):
        """
        creates the graphs mentioned in the graphNames list
        :param workbook: Excel workbook
        :param graphNames: list of graphs which need to be made
        :param sheetNames: list of sheet names
        :param hours: list of all the hours
        """
        wb = workbook

        pid = wb[graphNames[0]]
        ff = wb[graphNames[1]]
        voc = wb[graphNames[2]]
        isc = wb[graphNames[3]]

        graphs = [pid, ff, voc, isc]

        for j in range(0,len(graphs),1):
            chartObj = openpyxl.chart.ScatterChart()
            chartObj.legend.position = 'b'

            if j == 0:
                chartObj.x_axis.title = 'Time [h]'
                chartObj.y_axis.title = '%PID [%]'
                chartObj.y_axis.scaling.max = 100
                x = 1
                y = 17
            else:
                chartObj.x_axis.title = '%PIDs [%]'
                chartObj.x_axis.scaling.max = 100
                x = 17
                if j == 1:
                    chartObj.y_axis.title = 'Fill Factor [%]'
                    y = 8
                if j == 2:
                    chartObj.y_axis.title = 'Voc [mV]'
                    y = 7
                if j == 3:
                    chartObj.y_axis.title = 'Isc [mA]'
                    y = 5

            for i in range(0, len(sheetNames), 1):
                xvalues = openpyxl.chart.Reference(wb[sheetNames[i]], min_col=x, min_row=2, max_col=x, max_row=len(hours) + 1)
                yvalues = openpyxl.chart.Reference(wb[sheetNames[i]], min_col=y, min_row=2, max_col=y, max_row=len(hours) + 1)
                seriesObj = openpyxl.chart.Series(yvalues, xvalues, title=str(sheetNames[i]))
                seriesObj.marker = openpyxl.chart.marker.Marker('circle')
                chartObj.append(seriesObj)

            graphs[j].add_chart(chartObj)


    def makeChartPscSm(sheetname, workbook, times, frame):
        """
        creates the IV graphs of the sample
        :param sheetname: name of the sheet
        :param workbook: Excel workbook
        :param times: list of all the time frames in minutes
        """

        wb = workbook
        sheet = wb[sheetname]

        chartObj = openpyxl.chart.ScatterChart()
        chartObj.x_axis.title = 'Voltage [V]'
        chartObj.y_axis.title = 'Current [A]'
        chartObj.x_axis.scaling.min = 0
        chartObj.legend.position = 'b'

        chartObj.title = 'IV'
        location = 'C20'
        beginCol = 0

        if frame == 'Psc':
            chartObj.y_axis.scaling.max = 0
            for i in range(1, sheet.max_column, 1):
                head = sheet.cell(row=1, column=i).value
                if head is None:
                    beginCol = i
                    break

        for i in range(0, len(times)):
            xvalues = openpyxl.chart.Reference(sheet,min_col=beginCol+2+(i*3),min_row=3,max_col=beginCol+2+(i*3),max_row=sheet.max_row)
            yvalues = openpyxl.chart.Reference(sheet,min_col=beginCol+1+(i*3),min_row=3,max_col=beginCol+1+(i*3),max_row=sheet.max_row)
            # if frame == 'Psc':
            seriesObj = openpyxl.chart.Series(yvalues, xvalues, title=str(times[i]) + ' min')
            # else:
            #     seriesObj = openpyxl.chart.Series(yvalues, xvalues, title=str(times[i]) + ' h')
            chartObj.append(seriesObj)

        sheet.add_chart(chartObj, location)


    def makeSeperateGraphsPsc(workbook, sheetNames, times):
        """
        creates the graphs mentioned in the graphNames list
        :param workbook: Excel workbook
        :param graphNames: list of graphs which need to be made
        :param sheetNames: list of sheet names
        :param hours: list of all the hours
        """
        wb = workbook
        sheet = wb['%PID']

        chartObj = openpyxl.chart.ScatterChart()
        chartObj.legend.position = 'b'


        chartObj.x_axis.title = 'Time [min]'
        chartObj.y_axis.title = '%PID [%]'
        chartObj.y_axis.scaling.max = 100
        x = 1
        y = 6

        for i in range(0, len(sheetNames), 1):
            if sheetNames[i] in wb.sheetnames:
                xvalues = openpyxl.chart.Reference(wb[sheetNames[i]], min_col=x, min_row=2, max_col=x, max_row=len(times) + 1)
                yvalues = openpyxl.chart.Reference(wb[sheetNames[i]], min_col=y, min_row=2, max_col=y, max_row=len(times) + 1)
                seriesObj = openpyxl.chart.Series(yvalues, xvalues, title=str(sheetNames[i]))
                seriesObj.marker = openpyxl.chart.marker.Marker('circle')
                chartObj.append(seriesObj)

        sheet.add_chart(chartObj)

    def makeChartRsh(wb, modules):
        sheet = wb['Rsh']

        chartObj = openpyxl.chart.ScatterChart()
        chartObj.legend.position = 'b'

        chartObj.x_axis.title = 'Time [min]'
        chartObj.y_axis.title = 'Rsh [Ohm]'

        for m in range(0, len(modules)):
            x = 1+m*3
            y = x+1
            xvalues = openpyxl.chart.Reference(sheet, min_col=x, min_row=3, max_col=x,max_row=sheet.max_row)
            yvalues = openpyxl.chart.Reference(sheet, min_col=y, min_row=3, max_col=y, max_row=sheet.max_row)
            seriesObj = openpyxl.chart.Series(yvalues, xvalues, title=modules[m])
            chartObj.append(seriesObj)

        sheet.add_chart(chartObj)